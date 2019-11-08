# coding:utf-8

import openpyxl
import os

class  writeExcel():
    dir = 'common/testData'
    exec_dir=os.path.dirname(os.getcwd()) + "/" + dir + "/" +'data.xlsx'
    print('======',exec_dir)
     #打开一个将写的文件
    w = openpyxl.load_workbook(exec_dir)
    print('++++++++++',w)
    sheetname = w.sheetnames #在将写的文件创建sheet

    def writeResult(self,id,testdata,result):
        ws =self.w['asserSheet']
        #写入数据
        ws.cell(id+1,3).value=testdata
        ws.cell(id+1,4).value=result

        self.w.save(self.exec_dir + '\\' + 'data.xls')
